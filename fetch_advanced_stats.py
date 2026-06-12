import os
import json
import time
import requests
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------- Config ----------------
API_KEY = "neom-9172-sjou-mzxw"
BASE_URL = "https://api.boxtobox.ai/v1/wy/api/v3"

COMPETITION_WYID = 364
SEASON_ID = 191622
SLEEP_SECONDS = 0.15

EVENTS_DIR = Path(
    "/content/drive/MyDrive/Event data/Wyscout/Premier League/2025-2026"
)

OUT_XLSX = (
    "/content/drive/MyDrive/Event data/Wyscout/"
    "epl.xlsx"
)

COMPETITIONS_CSV = "/content/drive/MyDrive/Event data/Wyscout/domestic_competitions_all_areas.csv"


# ---------------- Helpers ----------------
def get_player_advanced_stats(player_id: int, comp_id: int, season_id: int | None):
    params = {"api_key": API_KEY, "compId": comp_id}
    if season_id is not None:
        params["seasonId"] = season_id

    r = requests.get(
        f"{BASE_URL}/players/{player_id}/advancedstats",
        params=params,
        timeout=30,
    )
    r.raise_for_status()
    return r.json()


def format_positions(position_list):
    if not isinstance(position_list, list):
        return None

    sorted_pos = sorted(position_list, key=lambda x: x.get("percent", 0), reverse=True)
    codes = [
        p.get("position", {}).get("code", "").upper()
        for p in sorted_pos
        if p.get("position", {}).get("code")
    ]
    return ",".join(codes) if codes else None


def build_player_lookup(events_dir: Path) -> pd.DataFrame:
    rows = []
    for fp in events_dir.glob("*.json"):
        try:
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)

            events = data.get("events", [])
            for ev in events:
                player = ev.get("player", {})
                team = ev.get("team", {})

                pid = player.get("id")
                if pid is None:
                    continue

                rows.append({
                    "playerId": int(pid),
                    "playerName": player.get("name"),
                    "teamName": team.get("name"),
                })
        except Exception:
            continue

    df = pd.DataFrame(rows)

    name_df = (
        df.groupby(["playerId", "playerName"])
        .size()
        .reset_index(name="n")
        .sort_values(["playerId", "n"], ascending=[True, False])
        .drop_duplicates("playerId")[["playerId", "playerName"]]
    )

    team_df = (
        df.groupby(["playerId", "teamName"])
        .size()
        .reset_index(name="n")
        .sort_values(["playerId", "n"], ascending=[True, False])
        .drop_duplicates("playerId")[["playerId", "teamName"]]
    )

    return name_df.merge(team_df, on="playerId", how="left")


def add_competition_name(df_final: pd.DataFrame, competitions_csv_path: str) -> pd.DataFrame:
    comps = pd.read_csv(competitions_csv_path)

    comp_id_col = None
    for c in ["competitionId", "wyId", "id"]:
        if c in comps.columns:
            comp_id_col = c
            break
    if comp_id_col is None:
        raise RuntimeError(
            f"Could not find a competition id column in competitions CSV. Columns: {comps.columns.tolist()}"
        )

    if "name" not in comps.columns:
        raise RuntimeError(
            f"Could not find 'name' column in competitions CSV. Columns: {comps.columns.tolist()}"
        )

    comps = comps.rename(columns={comp_id_col: "competitionId", "name": "competitionName"})
    comps["competitionId"] = pd.to_numeric(comps["competitionId"], errors="coerce")

    df_comp_col = None
    for c in ["competitionId", "competition.id", "compId"]:
        if c in df_final.columns:
            df_comp_col = c
            break

    if df_comp_col is None:
        df_final = df_final.copy()
        df_final["competitionId"] = COMPETITION_WYID
        df_comp_col = "competitionId"

    df_final = df_final.copy()
    df_final["competitionId"] = pd.to_numeric(df_final[df_comp_col], errors="coerce")

    extra = [c for c in ["area.name", "areaName", "country", "countryName"] if c in comps.columns]
    keep_cols = ["competitionId", "competitionName"] + extra
    comps_small = comps[keep_cols].drop_duplicates("competitionId")

    return df_final.merge(comps_small, on="competitionId", how="left")


# ---------------- Excel Formatting ----------------
# Colour palette
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
ID_BG       = "D6E4F0"   # light blue for identity cols
ALT_ROW_BG  = "F2F7FB"   # very faint blue stripe
BORDER_COL  = "BDD7EE"

# Columns considered "identity" (not stats)
IDENTITY_COLS = {"playerId", "playerName", "teamName", "positions_codes",
                 "competitionId", "competitionName"}

THIN = Side(style="thin", color=BORDER_COL)
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _make_header_style():
    return {
        "font": Font(name="Calibri", bold=True, color=HEADER_FG, size=11),
        "fill": PatternFill("solid", fgColor=HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": THIN_BORDER,
    }


def _make_id_style():
    return {
        "fill": PatternFill("solid", fgColor=ID_BG),
        "font": Font(name="Calibri", size=10),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": THIN_BORDER,
    }


def _make_data_style(alt: bool):
    bg = ALT_ROW_BG if alt else "FFFFFF"
    return {
        "fill": PatternFill("solid", fgColor=bg),
        "font": Font(name="Calibri", size=10),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": THIN_BORDER,
    }


def _apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)


def _col_width(series: pd.Series, header: str, max_width: int = 30) -> float:
    max_len = max(
        len(str(header)),
        series.dropna().astype(str).str.len().max() if not series.dropna().empty else 0,
    )
    return min(max_len + 2, max_width)


def _is_numeric_col(series: pd.Series) -> bool:
    return pd.api.types.is_numeric_dtype(series)


def _number_format(col_name: str, series: pd.Series) -> str:
    """Choose a sensible number format based on column name / dtype."""
    low = col_name.lower()
    if "id" in low:
        return "0"
    if any(k in low for k in ("percent", "pct", "rate", "%", "ratio", "avg", "average", "mean")):
        return "0.00"
    if _is_numeric_col(series):
        # integer-ish?
        if series.dropna().apply(float.is_integer).all() if not series.dropna().empty else True:
            return "0"
        return "0.00"
    return "General"


def format_stats_sheet(ws, df: pd.DataFrame, sheet_title: str = "Advanced Stats"):
    """Apply full formatting to a stats worksheet."""
    header_style = _make_header_style()
    id_style     = _make_id_style()

    cols = list(df.columns)
    n_rows, n_cols = df.shape

    # ---- Header row ----
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)

    # ---- Data rows ----
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        alt = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(cols, start=1):
            value = row[col_name]
            # convert numpy types
            if hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_name in IDENTITY_COLS:
                _apply_style(cell, id_style)
                if col_name == "playerName":
                    cell.alignment = Alignment(horizontal="left", vertical="center", bold=True)
            else:
                _apply_style(cell, _make_data_style(alt))
                if _is_numeric_col(df[col_name]):
                    cell.number_format = _number_format(col_name, df[col_name])

    # ---- Column widths ----
    for col_idx, col_name in enumerate(cols, start=1):
        width = _col_width(df[col_name], col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---- Freeze header + identity columns ----
    id_count = sum(1 for c in cols if c in IDENTITY_COLS)
    freeze_col = get_column_letter(id_count + 1)
    ws.freeze_panes = f"{freeze_col}2"

    # ---- Excel Table ----
    if n_rows > 0 and n_cols > 0:
        table_ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
        table = Table(displayName=sheet_title.replace(" ", "_"), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # ---- Row height ----
    ws.row_dimensions[1].height = 36  # header taller
    for r in range(2, n_rows + 2):
        ws.row_dimensions[r].height = 16

    ws.title = sheet_title


def format_failures_sheet(ws, df: pd.DataFrame):
    header_style = _make_header_style()
    ERR_BG = "FFF2CC"

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_style(cell, header_style)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            if hasattr(value, "item"):
                value = value.item()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=ERR_BG)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER

    for col_idx, col_name in enumerate(df.columns, start=1):
        width = _col_width(df[col_name], col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.title = "Failures"


# ---------------- Main ----------------
from google.colab import drive
drive.mount("/content/drive")

lookup = build_player_lookup(EVENTS_DIR)
player_ids = sorted(lookup["playerId"].unique())

rows = []
failures = []

for i, pid in enumerate(player_ids, 1):
    print(f"[{i}/{len(player_ids)}] player {pid}")
    try:
        data = get_player_advanced_stats(pid, comp_id=COMPETITION_WYID, season_id=SEASON_ID)

        if isinstance(data, dict) and "error" in data:
            failures.append({"playerId": pid, "error": data["error"]})
            continue

        flat = pd.json_normalize(data, sep=".")
        flat["playerId"] = pid
        flat["positions_codes"] = (
            format_positions(data.get("positions")) if isinstance(data, dict) else None
        )
        rows.append(flat)

    except Exception as e:
        failures.append({"playerId": pid, "error": str(e)})

    time.sleep(SLEEP_SECONDS)

df_stats = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
df_fail = pd.DataFrame(failures)

# Merge player name + team
df_final = lookup.merge(df_stats, on="playerId", how="left")

# Add competition name
df_final = add_competition_name(df_final, COMPETITIONS_CSV)

# Reorder: identity columns first, then all stats alphabetically
front = ["playerId", "playerName", "teamName", "positions_codes", "competitionId", "competitionName"]
front = [c for c in front if c in df_final.columns]
stat_cols = sorted([c for c in df_final.columns if c not in front])
df_final = df_final[front + stat_cols]

# ---- Write & format ----
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    df_final.to_excel(writer, sheet_name="advancedstats", index=False)
    if not df_fail.empty:
        df_fail.to_excel(writer, sheet_name="failures", index=False)

# Re-open and apply rich formatting
wb = load_workbook(OUT_XLSX)

ws_stats = wb["advancedstats"]
format_stats_sheet(ws_stats, df_final, sheet_title="Advanced Stats")

if not df_fail.empty and "failures" in wb.sheetnames:
    format_failures_sheet(wb["failures"], df_fail)

wb.save(OUT_XLSX)
print("Saved formatted Excel to:", OUT_XLSX)
df_final.head()
